"""Background processing worker.

Runs the full pipeline (read → AI extraction → docx generation → xlsx export)
in a daemon thread, posting progress and result messages to a ``queue.Queue``
that the main thread polls.

Message format
--------------
All messages are tuples whose first element is a ``str`` tag:

``("log", text, tag)``
    Append a line to the log panel.  ``tag`` is one of the colour tags
    registered on the CTkTextbox (``"success"``, ``"error"``, ``"warn"``,
    ``"dim"``, ``"accent"``, ``"bold"``), or an empty string for plain text.

``("progress", current, total)``
    Update the progress bar.

``("done", generated, errors, elapsed_seconds)``
    Processing finished successfully.

``("cancelled", done_so_far, total)``
    Processing was cancelled by the user.

``("error", message)``
    A fatal, unrecoverable error occurred.

Public exports:
    run_processing_worker: Start the worker thread.
"""

from __future__ import annotations

import os
import queue
import re
import sys
import threading
import time
from pathlib import Path
from typing import Any

from z7_officeletters.constants import MODELO_OFICIO, MODELO_REQUERIMENTO_PESAR, MODELO_PLANILHA, MODELO_ENVELOPE, ENDERECAMENTO_PADRAO, PASTA_SAIDA, PASTA_PLANILHA, PASTA_ENVELOPES, RE_PROPOSITURA_SPLIT, detectar_tipo_propositura, numero_propositura
from z7_officeletters.core import ai as _ai
from z7_officeletters.core import address_db as _addr_db
from z7_officeletters.core import authors as _authors
from z7_officeletters.core import documents as _docs
from z7_officeletters.core import files as _files
from z7_officeletters.core import recipients as _recipients
from z7_officeletters.core import verification as _verification
from z7_officeletters.core.api_key import salvar_api_key
from z7_officeletters.core.logging_setup import log_file_path, registrar_conferencia_ia

__all__ = ["run_processing_worker"]


def _normalizar_dest(nome: str) -> str:
    """Normalize a recipient name to a stable grouping key (uppercase, collapsed whitespace)."""
    return " ".join(nome.upper().split())


def _get_min_prop_number(grupo_items: list[tuple[dict, dict, dict]]) -> int:
    """Extract numbers from all items in the group and return the minimum."""
    nums = []
    for d_item, _, _ in grupo_items:
        num_str = d_item.get("numero_mocao") or ""
        m = re.search(r'\d+', num_str)
        if m:
            nums.append(int(m.group(0)))
    return min(nums) if nums else 0


def _ordenar_grupos(
    grupos: dict[tuple[str, str], list[tuple[dict, dict, dict]]]
) -> list[tuple[tuple[str, str], list[tuple[dict, dict, dict]]]]:
    """Sort groups: requerimento_pesar first, then mocao. Within type, sort by min number, then recipient name."""
    return sorted(
        grupos.items(),
        key=lambda x: (
            0 if x[0][0] == "requerimento_pesar" else 1,
            _get_min_prop_number(x[1]),
            x[0][1],
        )
    )


def _worker_main(
    inputs: dict[str, Any],
    q: "queue.Queue[tuple[Any, ...]]",
    cancel_event: threading.Event,
) -> None:
    """Main worker body — executed in a daemon thread."""
    try:
        from google import genai  # noqa: PLC0415
        from docxtpl import DocxTemplate  # noqa: PLC0415
        from openpyxl import Workbook, load_workbook  # noqa: PLC0415

        q.put(("log", f"📋  Log: {log_file_path}", "dim"))

        salvar_api_key(inputs["api_key"])
        cliente = genai.Client(api_key=inputs["api_key"])

        model_input_limit = 0
        try:
            _model_info = cliente.models.get(model=_ai.MODELO_IA)
            model_input_limit = int(_model_info.input_token_limit or 0)
        except Exception:  # noqa: BLE001
            pass

        arquivos_proc: list[str] = inputs["arquivos"]
        todos_textos: list[tuple[str, str]] = []
        for arq in arquivos_proc:
            q.put(("log", f"📂  Lendo: {Path(arq).name}", "accent"))
            conteudo = _files.ler_arquivo_mocoes(arq)
            textos_arq = RE_PROPOSITURA_SPLIT.split(conteudo)
            for t in textos_arq:
                t = t.strip()
                if t and RE_PROPOSITURA_SPLIT.match(t):
                    todos_textos.append((t, detectar_tipo_propositura(t)))

        proposituras = sorted(
            todos_textos,
            key=lambda item: (0 if item[1] == "requerimento_pesar" else 1, numero_propositura(item[0])),
        )
        total = len(proposituras)
        n_mocoes = sum(1 for _, tp in proposituras if tp == "mocao")
        n_pesar = sum(1 for _, tp in proposituras if tp == "requerimento_pesar")
        partes = []
        if n_mocoes:
            partes.append(f"{n_mocoes} moção(oes)")
        if n_pesar:
            partes.append(f"{n_pesar} requerimento(s) de pesar")
        resumo = " e ".join(partes) if partes else f"{total} propositura(s)"
        q.put(("log", f"\n❆  {resumo} encontrada(s). Iniciando IA…\n", "bold"))
        q.put(("progress", 0, total))

        Path(PASTA_SAIDA).mkdir(parents=True, exist_ok=True)

        _app_root = (
            Path(sys.executable).parent
            if getattr(sys, "frozen", False)
            else Path(__file__).parent.parent.parent.parent.parent
        )
        _meipass = Path(getattr(sys, "_MEIPASS", ""))

        def _resolve_template(rel: str) -> Path:
            p = _app_root / rel
            if not p.exists() and getattr(sys, "frozen", False):
                p = _meipass / rel
            return p

        modelo_oficio = _resolve_template(MODELO_OFICIO)
        modelo_requerimento_pesar = _resolve_template(MODELO_REQUERIMENTO_PESAR)
        modelo_envelope = _resolve_template(MODELO_ENVELOPE)

        # Address DB — optional; the app degrades gracefully when absent.
        _db_path = _resolve_template(ENDERECAMENTO_PADRAO)
        if _db_path.exists():
            q.put(("log", f"📒  Base de endereçamentos: {_db_path.name}", "dim"))
            _addr_db.buscar_endereco("__warmup__", db_path=_db_path)  # prime cache
        else:
            _db_path = None  # type: ignore[assignment]
            q.put(("log", "  ⚠  enderecamentos_padrao.docx não encontrado — usando apenas dados da IA.", "warn"))

        if not modelo_oficio.exists():
            q.put(("error", f"Arquivo 'modelo_mocao.docx' não encontrado.\n{modelo_oficio}"))
            return

        if not modelo_envelope.exists():
            try:
                _docs.criar_modelo_envelope(modelo_envelope)
            except Exception as exc:
                q.put(("log", f"  ⚠  Não foi possível criar o template de envelope: {exc}", "warn"))

        dados_planilha: list[list[str]] = []
        numero_atual: int = inputs["num_inicial"]
        year: int = int(inputs["data_iso"][:4])
        erros = 0
        inicio = time.time()
        total_prompt_tokens = 0
        total_candidates_tokens = 0
        total_tokens = 0
        registros_verificacao: list[_verification.RegistroOficio] = []

        # ── Phase 1: AI extraction ────────────────────────────────────────────
        # Process each propositura individually; collect validated data dicts.
        extracted: list[tuple[str, dict]] = []  # (tipo_propositura, dados)

        for i, (texto, tipo_propositura) in enumerate(proposituras, 1):
            if cancel_event.is_set():
                q.put(("cancelled", i - 1, total, "proposituras"))
                return

            _tipo_label = "Moção" if tipo_propositura == "mocao" else "Req. Pesar"
            q.put(("log", f"\n▶  Propositura {i}/{total}  ·  {_tipo_label}", "accent"))
            q.put(("progress", i - 1, total))

            try:
                dados = _ai.extrair_dados_com_ia(
                    texto, cliente,
                    tipo_propositura=tipo_propositura,
                    cancel_event=cancel_event,
                    on_rate_limit=lambda msg: q.put(("log", f"  {msg}", "warn")),
                    instrucoes_complementares=inputs.get("instrucoes_complementares"),
                )
            except RuntimeError as exc:
                if cancel_event.is_set():
                    q.put(("cancelled", i - 1, total, "proposituras"))
                    return
                q.put(("log", f"  ✖  Erro: {exc}", "error"))
                erros += 1
                continue
            except Exception as exc:  # noqa: BLE001
                q.put(("log", f"  ✖  Erro: {exc}", "error"))
                erros += 1
                continue

            usage = dados.pop("_usage", {"prompt_tokens": 0, "candidates_tokens": 0, "total_tokens": 0})
            alertas_ia = dados.pop("_alertas", [])
            total_prompt_tokens += usage["prompt_tokens"]
            total_candidates_tokens += usage["candidates_tokens"]
            total_tokens += usage["total_tokens"]
            if usage["total_tokens"]:
                _saldo_str = (
                    f"  |  saldo: {(model_input_limit - usage['prompt_tokens']):,}"
                    if model_input_limit else ""
                )
                q.put(("log",
                    f"  🔢  Tokens: {usage['total_tokens']:,} "
                    f"(entrada: {usage['prompt_tokens']:,} | saída: {usage['candidates_tokens']:,}){_saldo_str}",
                    "dim"))
            for alerta in alertas_ia:
                q.put(("log", f"  ℹ  {alerta}", "dim"))

            # Normalise motion/requerimento number to just the numeric part.
            num_raw = dados.get("numero_requerimento") or dados.get("numero_mocao", "")
            dados["numero_mocao"] = _docs.normalizar_numero_mocao(str(num_raw))
            extracted.append((tipo_propositura, dados))

            _num_extr = dados.get("numero_mocao") or "–"
            _tipo_extr = dados.get("tipo_mocao", "")
            _falecido_extr = dados.get("falecido", "")
            _dests_extr = [d.get("nome", "?") for d in dados.get("destinatarios", [])]
            _label_extr = _tipo_extr or ("Pesar" if tipo_propositura == "requerimento_pesar" else "")
            _dests_str = " / ".join(_dests_extr) if _dests_extr else "sem destinatários"
            _sum_label = f"nº {_num_extr}"
            if _label_extr:
                _sum_label += f" ({_label_extr})"
            if _falecido_extr:
                _sum_label += f" · {_falecido_extr}"
            q.put(("log", f"  ↳  {_sum_label}  →  {_dests_str}", "dim"))

        # ── Phase 2: group by (tipo_propositura, recipient) ───────────────────
        # When multiple propositions share the same recipient, they are merged
        # into a single office letter so that one ofício covers all of them.
        #
        # Key:   (tipo_propositura, normalized_dest_name)
        # Value: list of (dados, raw_dest_dict, processed_info) triples
        grupos: dict[tuple[str, str], list[tuple[dict, dict, dict]]] = {}

        for tipo_propositura, dados in extracted:
            for dest in dados["destinatarios"]:
                # Enrich recipient data: DB (priority 1) > propositura (priority 2)
                dest_proc = dict(dest)  # shallow copy to avoid mutating AI data
                db_entry = _addr_db.buscar_endereco(dest["nome"], db_path=_db_path)
                if db_entry:
                    # DB is the most authoritative source — override all fields it supplies.
                    dest_proc["nome"] = db_entry.nome
                    if db_entry.cargo:
                        dest_proc["cargo_ou_tratamento"] = db_entry.cargo
                    if db_entry.endereco:
                        dest_proc["endereco"] = db_entry.endereco
                    if db_entry.email:
                        dest_proc["email"] = db_entry.email

                info = _recipients.processar_destinatario(dest_proc)

                # Override honorifics when DB supplies a richer tratamento string.
                if db_entry:
                    _recipients.aplicar_tratamento_db(info, db_entry.tratamento)

                dest_key = (tipo_propositura, _normalizar_dest(dest_proc["nome"]))
                if dest_key not in grupos:
                    grupos[dest_key] = []
                grupos[dest_key].append((dados, dest, info))

        n_grupos = len(grupos)
        q.put(("progress", total, total))
        if n_grupos:
            q.put(("log", f"\n📬  {n_grupos} ofício(s) a gerar após agrupamento...\n", "bold"))

        unrecognized_authors: set[str] = set()

        # ── Phase 3: generate one letter per group ────────────────────────────
        grupos_ordenados = _ordenar_grupos(grupos)
        for dest_key, grupo in grupos_ordenados:
            if cancel_event.is_set():
                q.put(("cancelled", len(dados_planilha), n_grupos, "ofícios"))
                return

            tipo_propositura = dest_key[0]
            n_props = len(grupo)

            # Merge proposition data from every item in the group.
            all_autores: list[str] = []
            for d_item, _dest_raw, _info_item in grupo:
                for a in d_item["autores"]:
                    if a not in all_autores:
                        all_autores.append(a)

            # Warn user if any extracted author name is not mapped in config.json
            for a in all_autores:
                if _authors.sigla_autor(a) == "indef":
                    unrecognized_authors.add(a)
                    q.put((
                        "log",
                        f"  ⚠  Autor '{a}' não encontrado no config.json. Usando sigla 'indef'.\n"
                        f"      Para corrigir, adicione-o via Editor de Configurações (Avançado) ou no 'config.json'.",
                        "warn"
                    ))

            texto_autoria, sigla_autores = _authors.formatar_autores(all_autores)

            nums_mocao = [d_item["numero_mocao"] for d_item, _, __ in grupo]
            num_mocao_merged = _docs.formatar_lista_pt(nums_mocao)

            tipos_mocao = [
                str(d_item.get("tipo_mocao", ""))
                for d_item, _, __ in grupo
                if d_item.get("tipo_mocao")
            ]
            tipo_mocao_merged = _docs.formatar_lista_pt(tipos_mocao) if tipos_mocao else ""

            falecidos = [
                str(d_item.get("falecido", ""))
                for d_item, _, __ in grupo
                if d_item.get("falecido")
            ]
            falecido_merged = _docs.formatar_lista_pt(falecidos) if falecidos else ""

            # Recipient info is identical for all items in the group
            # (they share the same destination); use the first entry.
            _dados0, dest0, info = grupo[0]

            _tipo_ofc = (
                "Req. Pesar"
                if tipo_propositura == "requerimento_pesar"
                else (f"Moção de {tipo_mocao_merged}" if tipo_mocao_merged else "Moção")
            )
            _agrup_label = f"  ({n_props}×)" if n_props > 1 else ""
            num_str = f"{numero_atual:03d}"
            q.put(("log",
                f"\n  📄  Ofício {num_str}{_agrup_label}  ·  {_tipo_ofc}  →  {dest_key[1]}",
                "bold"))
            _autores_log = "  /  ".join(all_autores) if all_autores else "—"
            q.put(("log", f"      {_autores_log}", "dim"))

            sigla_redator = inputs["sigla"]

            # Plural-aware phrase fragments used by both templates.
            # These allow the assunto line and body paragraph to read correctly
            # regardless of whether one or many propositions are grouped.
            _designacao_prop, _copia_art, _aprovada_s = _docs.frases_propositura(
                tipo_propositura, tipo_mocao_merged, n_props
            )

            ctx: dict[str, str] = {
                "num_oficio":            num_str,
                "data_extenso":          inputs["data_extenso"],
                "tipo_mocao":            tipo_mocao_merged,
                "num_mocao":             num_mocao_merged,
                "falecido":              falecido_merged,
                "tipo_propositura":      tipo_propositura,
                "sigla_redator":         sigla_redator,
                "vocativo":              info["vocativo"],
                "pronome_corpo":         info["pronome_corpo"],
                "texto_autoria":         texto_autoria,
                "tratamento_rodape":     info["tratamento_rodape"],
                "destinatario_nome":     info["destinatario_nome"],
                "destinatario_endereco": info["destinatario_endereco"],
                "designacao_propositura": _designacao_prop,
                "copia_art":             _copia_art,
                "aprovada_s":            _aprovada_s,
                # Uppercase aliases for Word template placeholders
                "NUM_OFICIO":            num_str,
                "DATA_EXTENSO":          inputs["data_extenso"],
                "TIPO_MOCAO":            tipo_mocao_merged,
                "NUM_MOCAO":             num_mocao_merged,
                "FALECIDO":              falecido_merged,
                "TIPO_PROPOSITURA":      tipo_propositura,
                "SIGLA_REDATOR":         sigla_redator,
                "VOCATIVO":              info["vocativo"],
                "PRONOME_CORPO":         info["pronome_corpo"],
                "TEXTO_AUTORIA":         texto_autoria,
                "TRATAMENTO_RODAPE":     info["tratamento_rodape"],
                "DESTINATARIO_NOME":     info["destinatario_nome"],
                "DESTINATARIO_ENDERECO": info["destinatario_endereco"],
                "DESIGNACAO_PROPOSITURA": _designacao_prop,
                "COPIA_ART":             _copia_art,
                "APROVADA_S":            _aprovada_s,
            }

            if tipo_propositura == "requerimento_pesar":
                ctx["vocativo"]           = "Ilustríssimos Senhores(as)"
                ctx["VOCATIVO"]           = "Ilustríssimos Senhores(as)"
                ctx["pronome_corpo"]      = "Vossas Senhorias"
                ctx["PRONOME_CORPO"]      = "Vossas Senhorias"
                ctx["tratamento_rodape"]  = "Aos familiares do Sr.(ª),"
                ctx["TRATAMENTO_RODAPE"]  = "Aos familiares do Sr.(ª),"
                ctx["destinatario_nome"]  = falecido_merged.upper()
                ctx["DESTINATARIO_NOME"]  = falecido_merged.upper()

                _tmpl = modelo_requerimento_pesar
                if not _tmpl.exists():
                    q.put(("log",
                        f"  ⚠  Template 'modelo_requer_pesar.docx' não encontrado — "
                        f"usando modelo_mocao.docx como fallback.",
                        "warn"))
                    _tmpl = modelo_oficio
            else:
                _tmpl = modelo_oficio

            doc = DocxTemplate(str(_tmpl))
            doc.render(ctx)

            nome = _docs.construir_nome_arquivo(
                num_str,
                inputs["sigla"],
                tipo_mocao_merged,
                num_mocao_merged,
                info["envio"],
                dest0["nome"],
                sigla_autores,
                ano=year,
                tipo_propositura=tipo_propositura,
            )
            caminho_oficio = os.path.join(PASTA_SAIDA, nome)
            doc.save(caminho_oficio)
            q.put(("log", f"  ✔  {nome}", "success"))

            # Generate envelope if delivery method is "Carta"
            if info["envio"] == "Carta":
                Path(PASTA_ENVELOPES).mkdir(parents=True, exist_ok=True)
                if modelo_envelope.exists():
                    try:
                        doc_env = DocxTemplate(str(modelo_envelope))
                        doc_env.render(ctx)
                        nome_dest_safe = _docs._RE_NOME_INVALIDO.sub("", _docs._titlecase_nome(dest0["nome"]))
                        nome_envelope = f"Envelope - Of. {num_str} - {nome_dest_safe}.docx"
                        caminho_envelope = os.path.join(PASTA_ENVELOPES, nome_envelope)
                        doc_env.save(caminho_envelope)
                        q.put(("log", f"  ✉  Envelope gerado: {nome_envelope}", "success"))
                    except Exception as exc:
                        q.put(("log", f"  ✖  Erro ao gerar envelope para {dest0['nome']}: {exc}", "error"))
                        erros += 1
                else:
                    q.put(("log", "  ⚠  Template de envelope não encontrado — ignorando geração do envelope.", "warn"))

            if tipo_propositura == "requerimento_pesar":
                plural_s = "s" if n_props > 1 else ""
                assunto = f"Encaminha Requerimento{plural_s} de Pesar nº {num_mocao_merged}/{year}"
            else:
                plural_oes = "ções" if n_props > 1 else "ção"
                assunto = f"Encaminha Mo{plural_oes} de {tipo_mocao_merged} nº {num_mocao_merged}/{year}"

            dados_planilha.append([
                num_str,
                inputs["data_iso"],
                f"{info['tratamento_rodape']} {_docs._titlecase_nome(info['destinatario_nome'])}".strip(),
                assunto,
                ", ".join(
                    f"{_docs._titlecase_nome(_authors._resolve_casing(a.lower(), _authors.norm(a), a))} ({_authors.sigla_autor(a)})"
                    for a in all_autores
                ),
                info["envio"],
                inputs["sigla"],
            ])

            # Register this letter for Phase 6 verification.
            registros_verificacao.append(_verification.RegistroOficio(
                caminho=caminho_oficio,
                nome_arquivo=nome,
                ctx=dict(ctx),
                dados_grupo=[d for d, _, __ in grupo],
                dest_raw=dict(dest0),
                info=dict(info),
                n_props=n_props,
                tipo_propositura=tipo_propositura,
                template_path=str(_tmpl),
                linha_planilha_idx=len(dados_planilha) - 1,
            ))

            numero_atual += 1

        # ── Phase 5: Verification (conference) ───────────────────────────────
        # Runs before the spreadsheet is written to disk so that any corrections
        # to dados_planilha rows are reflected in the final .xlsx file.
        if registros_verificacao:
            relatorio_conf = _verification.conferir_trabalho(
                registros_verificacao, dados_planilha, q
            )
            registrar_conferencia_ia(relatorio_conf)
            erros += relatorio_conf.total_incorrigiveis

        # ── Phase 6: Excel spreadsheet ────────────────────────────────────────
        q.put(("log", "\n📊  Gerando planilha Excel…", "accent"))
        if getattr(sys, "frozen", False):
            modelo_xlsx = Path(sys.executable).parent / MODELO_PLANILHA
        else:
            modelo_xlsx = Path(__file__).parent.parent.parent.parent.parent / MODELO_PLANILHA

        if modelo_xlsx.exists():
            wb = load_workbook(str(modelo_xlsx))
            ws = wb.active
            assert ws is not None
        else:
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(["Of. n.º", "Data", "Destinatário", "Assunto", "Vereador", "Envio", "Autor"])

        ws.title = f"Controle {year}"
        for row in dados_planilha:
            ws.append(row)

        Path(PASTA_PLANILHA).mkdir(parents=True, exist_ok=True)
        wb.save(os.path.join(PASTA_PLANILHA, "CONTROLE_OFICIOS.xlsx"))

        elapsed = time.time() - inicio
        if total_tokens:
            q.put(("log",
                f"\n🔢  Tokens consumidos: {total_tokens:,} total "
                f"(entrada: {total_prompt_tokens:,} | saída: {total_candidates_tokens:,})",
                "accent"))
        
        if unrecognized_authors:
            autores_list = "\n".join(f"      • {a}" for a in sorted(unrecognized_authors))
            q.put((
                "log",
                f"\n⚠  ATENÇÃO: Os seguintes autores não foram identificados no config.json (sigla 'indef'):\n"
                f"{autores_list}\n"
                f"   Para corrigir, adicione-os via Editor de Configurações (Avançado) ou edite o 'config.json'.\n",
                "warn"
            ))

        q.put(("done", len(dados_planilha), erros, elapsed, total_tokens))

    except Exception as exc:  # noqa: BLE001
        q.put(("error", str(exc)))


def run_processing_worker(
    inputs: dict[str, Any],
    q: "queue.Queue[tuple[Any, ...]]",
    cancel_event: threading.Event,
) -> threading.Thread:
    """Start the processing worker in a background daemon thread.

    Args:
        inputs: Processing parameters with keys ``num_inicial``, ``sigla``,
            ``data_extenso``, ``data_iso``, ``arquivos``, and ``api_key``.
        q: Queue to post progress/result messages on.
        cancel_event: Event that, when set, requests graceful cancellation.

    Returns:
        The started ``Thread`` instance.
    """
    t = threading.Thread(
        target=_worker_main,
        args=(inputs, q, cancel_event),
        daemon=True,
    )
    t.start()
    return t
